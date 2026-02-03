import os
import requests
import pandas as pd
import joblib
import xlsxwriter
from pathlib import Path
from openpyxl import load_workbook
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
solcast_api_key = os.getenv("SOLCAST_API_KEY") or os.getenv("solcast_api_key")

# Get the directory where this script is located
BASE_DIR = Path(__file__).parent
ASTRO_DIR = BASE_DIR / "Astro"
SOLCAST_DIR = ASTRO_DIR / "Solcast"

# Create directories if they don't exist
ASTRO_DIR.mkdir(exist_ok=True)
SOLCAST_DIR.mkdir(exist_ok=True)


def fetching_Astro_data_15min():
	"""
	Fetch 15-minute solar forecast data from Solcast API.

	Returns:
		pd.DataFrame: The fetched weather/solar data
	"""
	lat = 46.971281
	lon = 23.674705
	# Fetch data from the API
	api_url = "https://api.solcast.com.au/data/forecast/radiation_and_weather?latitude={}&longitude={}&hours=168&output_parameters=air_temp,ghi,azimuth,cloud_opacity,dewpoint_temp,relative_humidity,zenith&period=PT15M&format=csv&time_zone=3&api_key={}".format(lat, lon, solcast_api_key)
	response = requests.get(api_url)
	print("Fetching data...")
	csv_path = SOLCAST_DIR / "Luna_15min.csv"
	if response.status_code == 200:
		# Write the content to a CSV file
		with open(csv_path, 'wb') as file:
			file.write(response.content)
	else:
		print(response.text)  # Add this line to see the error message returned by the API
		raise Exception(f"Failed to fetch data: Status code {response.status_code}")
	# Adjusting the values to EET time
	data = pd.read_csv(csv_path)
	return data

def predicting_exporting_Astro_15min(interval_from, interval_to, limitation_percentage):
	# Creating the forecast_dataset df
	csv_path = SOLCAST_DIR / "Luna_15min.csv"
	df = pd.read_csv(csv_path)
	# Convert the 'period_end' column to datetime, handling errors
	df['period_end'] = pd.to_datetime(df['period_end'], errors='coerce', format='%Y-%m-%dT%H:%M:%SZ')

	# Drop any rows with NaT in 'period_end'
	df.dropna(subset=['period_end'], inplace=True)

	# Shift the 'period_end' column by 2 hours
	df['period_end'] = df['period_end'] + pd.Timedelta(hours=2)

	# Creating the Interval column
	df['Interval'] = df.period_end.dt.hour * 4 + df.period_end.dt.minute // 15 + 1

	df.rename(columns={'period_end': 'Data', 'ghi': 'Radiatie', "air_temp": "Temperatura", "cloud_opacity": "Nori", "azimuth": "Azimuth", "zenith": "Zenith", "dewpoint_temp": "Dewpoint", "relative_humidity": "Umiditate"}, inplace=True)

	df = df[["Data", "Interval", "Temperatura", "Nori", "Radiatie", "Dewpoint", "Zenith", "Azimuth", "Umiditate"]]

	model_path = ASTRO_DIR / "rs_xgb_Astro_prod_15min_0325.pkl"
	xgb_loaded = joblib.load(model_path)

	df["Month"] = df.Data.dt.month
	dataset = df.copy()
	forecast_dataset = dataset[["Interval", "Temperatura", "Nori", "Radiatie", "Month"]]

	preds = xgb_loaded.predict(forecast_dataset.values)
	
	# Rounding each value in the list to the third decimal
	rounded_values = [round(value, 3) for value in preds]
	
	#Exporting Results to Excel
	results_path = ASTRO_DIR / "Results_Production_Astro_xgb_15min.xlsx"
	workbook = xlsxwriter.Workbook(str(results_path))
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value, interval in zip(rounded_values, dataset.Interval):
		if interval_from * 4 <= interval <= interval_to * 4:
			worksheet.write(row, col + 2, value * (1 - limitation_percentage / 100), decimal_format)
			row += 1
		else:
			worksheet.write(row, col + 2, value, decimal_format)
			row += 1

	row = 1
	for Data, Interval in zip(dataset.Data, dataset.Interval):
		worksheet.write(row, col + 0, Data, date_format)
		worksheet.write(row, col + 1, Interval)
		row += 1
	workbook.close()
	# Formatting the Results file
	# Step 1: Open the Excel file
	file_path = str(results_path)
	workbook = load_workbook(filename=file_path)
	worksheet = workbook['Production_Predictions']  # Adjust the sheet name as necessary

	# Step 2: Directly round the values in column C and write them back
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value is not None:  # Check if the cell is not empty
			# Round the value to 3 decimal places and write it back to column C
			worksheet.cell(row, 3).value = round(original_value, 3)
		
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value < 0.01:  # Check if the value is less than 0.01
			# Residual values are rounded to 0.000
			worksheet.cell(row, 3).value = 0
	# Save the workbook with the rounded values
	workbook.save(filename=file_path)
	workbook.close()
	# Open the existing workbook
	# Load the Excel file into a DataFrame
	df = pd.read_excel(file_path)
	
	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])
	
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Interval"].astype(str)
	df.to_excel(file_path, index=False)
	return dataset